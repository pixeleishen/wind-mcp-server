"""Read codes and time ranges from an Excel file.

Expected Excel format:
  Column A (required): code     — e.g. 000001.SZ
  Column B (optional): beginTime — e.g. 2025-01-01
  Column C (optional): endTime   — e.g. 2025-01-31

Header row is auto-detected: if the first row contains known header names
(code, beginTime, endTime) it is skipped; otherwise all rows are treated as data.
"""

from openpyxl import load_workbook

HEADER_ALIASES = {
    "code": "code",
    "codes": "code",
    "股票代码": "code",
    "证券代码": "code",
    "wind代码": "code",
    "begintime": "beginTime",
    "begin_time": "beginTime",
    "start": "beginTime",
    "startdate": "beginTime",
    "start_date": "beginTime",
    "开始日期": "beginTime",
    "endtime": "endTime",
    "end_time": "endTime",
    "end": "endTime",
    "enddate": "endTime",
    "end_date": "endTime",
    "结束日期": "endTime",
}


def _normalize_header(val: str) -> str | None:
    return HEADER_ALIASES.get(val.strip().lower().replace(" ", ""))


def read_excel(path: str) -> dict:
    """Return {"codes": "000001.SZ,000002.SZ", "beginTime": ..., "endTime": ...}."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError(f"Excel file is empty: {path}")

    # Detect header row
    col_map = {}  # column index -> field name
    data_start = 0

    first_row = rows[0]
    for idx, cell in enumerate(first_row):
        if cell is not None:
            field = _normalize_header(str(cell))
            if field:
                col_map[idx] = field

    if col_map:
        data_start = 1  # skip header
    else:
        # No header detected — assume column order: code, beginTime, endTime
        col_map = {0: "code"}
        if len(first_row) > 1:
            col_map[1] = "beginTime"
        if len(first_row) > 2:
            col_map[2] = "endTime"

    codes = []
    begin_times = []
    end_times = []

    for row in rows[data_start:]:
        for idx, field in col_map.items():
            if idx >= len(row) or row[idx] is None:
                continue
            val = str(row[idx]).strip()
            if not val:
                continue
            if field == "code":
                codes.append(val)
            elif field == "beginTime":
                begin_times.append(val)
            elif field == "endTime":
                end_times.append(val)

    if not codes:
        raise ValueError(f"No codes found in Excel: {path}")

    result = {"codes": ",".join(codes)}

    if begin_times:
        result["beginTime"] = min(begin_times)
    if end_times:
        result["endTime"] = max(end_times)

    return result
